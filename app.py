import streamlit as st
import pandas as pd
from dataclasses import asdict
from main import GoogleMapsScraper
from locations import AREA_MAPPINGS
import time
import asyncio
import sys
import base64
import os
import io
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import threading
import logging

# Fix for Windows asyncio loop policy
if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

st.set_page_config(page_title="Beyond Ads Scraper", page_icon="🚀", layout="wide")

# -------------------------------------------------------
# Scraper Thread Function
# -------------------------------------------------------
def run_scraper_thread(search_query, total_target, required_area, excluded_areas, allowed_areas, results_list, stop_event, status_dict):
    try:
        status_dict["text"] = "Initializing browser..."
        scraper = GoogleMapsScraper()
        
        status_dict["text"] = "Navigating to Google Maps..."
        success = scraper.start(search_query, total_target, required_area, excluded_areas, allowed_areas)
        
        if not success:
            status_dict["text"] = "Failed to find results (Timeout or Blocking)."
            status_dict["error"] = True
            return

        status_dict["text"] = "Scraping in progress..."
        while len(scraper.places) < total_target and not stop_event.is_set():
            # Perform one step
            item = scraper.step(should_stop_callback=stop_event.is_set)
            
            if item:
                # Thread-safe append (lists are thread-safe in CPython for append)
                results_list.append(asdict(item))
                status_dict["text"] = f"Found: {item.name}"
            
            # Check if scraper is stuck or finished (replicating logic from main.py)
            # Also check stop event here just in case
            if stop_event.is_set():
                break
            listings_locator = scraper.page.locator('//a[contains(@href, "https://www.google.com/maps/place")]')
            try:
                listings_count = listings_locator.count()
                if scraper.processed_count >= listings_count:
                    # Try scroll
                    status_dict["text"] = "Scrolling for more results..."
                    scraper.page.mouse.wheel(0, 10000)
                    scraper.page.wait_for_timeout(2000)
                    if scraper.page.locator('//a[contains(@href, "https://www.google.com/maps/place")]').count() <= listings_count:
                        # No new results after scroll
                        status_dict["text"] = "No more results found."
                        break
            except:
                break
                
            time.sleep(0.1)

        scraper.stop()
        status_dict["text"] = "Finished."
    except Exception as e:
        status_dict["text"] = f"Error: {str(e)}"
        status_dict["error"] = True
        logging.error(f"Thread error: {e}")

# -------------------------------------------------------
# Session State Initialization
# -------------------------------------------------------
if "results" not in st.session_state:
    st.session_state.results = []
if "is_scraping" not in st.session_state:
    st.session_state.is_scraping = False
if "start_time" not in st.session_state:
    st.session_state.start_time = 0
if "search_query" not in st.session_state:
    st.session_state.search_query = ""
if "stop_event" not in st.session_state:
    st.session_state.stop_event = None
if "scraper_thread" not in st.session_state:
    st.session_state.scraper_thread = None
if "status_dict" not in st.session_state:
    st.session_state.status_dict = {"text": "", "error": False}

# Cleanup legacy session state
if "scraper" in st.session_state:
    del st.session_state.scraper
if "is_paused" in st.session_state:
    del st.session_state.is_paused

# -------------------------------------------------------
# Utility Functions
# -------------------------------------------------------
def get_base64_of_bin_file(bin_file):
    with open(bin_file, 'rb') as f:
        data = f.read()
    return base64.b64encode(data).decode()

def set_png_as_page_bg(png_file):
    if not os.path.exists(png_file):
        return
    
    bin_str = get_base64_of_bin_file(png_file)
    mime_type = "image/png"
    if png_file.lower().endswith(".jpg") or png_file.lower().endswith(".jpeg"):
        mime_type = "image/jpeg"
        
    page_bg_img = '''
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Poppins', sans-serif; color: #ffffff; }
    .stApp {
        background-image: url("data:%s;base64,%s");
        background-size: cover;
        background-position: center;
        background-attachment: fixed;
    }
    
    /* Input Styling - Zero Transparency */
    div[data-baseweb="input"], div[data-baseweb="base-input"], div[data-baseweb="select"] > div,
    .stTextInput > div > div, .stSelectbox > div > div, .stNumberInput > div > div {
        background-color: #000000 !important;
        border: 1px solid #444444 !important;
        border-radius: 8px !important;
        color: #ffffff !important;
    }
    input[type="text"], input[type="number"], .stTextInput input, .stNumberInput input {
        background-color: #000000 !important;
        color: #ffffff !important;
        caret-color: #ffffff !important;
    }
    div[data-baseweb="input"]:focus-within, div[data-baseweb="select"]:focus-within {
        border-color: #ffffff !important;
        box-shadow: 0 0 8px rgba(255, 255, 255, 0.2) !important;
    }
    div[data-baseweb="popover"], div[data-baseweb="menu"], ul[data-baseweb="menu"] {
        background-color: #000000 !important;
        border: 1px solid #333333 !important;
    }
    li[data-baseweb="option"] { color: #cccccc !important; }
    li[data-baseweb="option"]:hover, li[data-baseweb="option"][aria-selected="true"] {
        background-color: #222222 !important;
        color: #ffffff !important;
    }
    div[data-baseweb="select"] span { color: #ffffff !important; }
    
    /* Text & Headers */
    h1, h2, h3, h4, h5, h6, p, label, span, .stMarkdown {
        color: #ffffff !important;
        text-shadow: 0px 2px 4px rgba(0,0,0,0.9);
    }
    [data-testid="stMetricValue"] { color: #ffffff !important; font-weight: 700 !important; text-shadow: 0 2px 4px rgba(0,0,0,0.9); }
    [data-testid="stMetricLabel"] { color: #dddddd !important; }

    /* Buttons */
    .stButton > button {
        background-color: #000000 !important;
        color: #ffffff !important;
        border: 1px solid #ffffff !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        text-transform: uppercase;
        letter-spacing: 1px;
        transition: all 0.2s ease;
    }
    .stButton > button:hover {
        background-color: #ffffff !important;
        color: #000000 !important;
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.5);
    }
    svg { fill: #ffffff !important; }
    .stProgress > div > div > div > div { background-color: #ffffff !important; }
    
    /* Mobile Optimization */
    @media only screen and (max-width: 600px) {
        /* Buttons - Make them larger and full width */
        .stButton > button, .stDownloadButton > button {
            width: 100%% !important;
            padding: 15px !important;
            font-size: 1.1em !important;
            height: auto !important;
            min-height: 50px !important;
            margin-bottom: 10px !important;
        }
        
        /* Inputs - Increase touch area */
        div[data-baseweb="input"] input, div[data-baseweb="select"] {
            font-size: 16px !important; /* Prevent zoom on iOS */
            min-height: 45px !important;
        }
        
        /* Metrics - Stack them nicely */
        [data-testid="stMetricValue"] {
            font-size: 1.5rem !important;
        }
        
        /* Headers - Adjust size */
        h1 { font-size: 2rem !important; }
        
        /* Ensure columns have spacing when stacked */
        [data-testid="column"] {
            margin-bottom: 15px !important;
            min-width: 100%% !important; /* Force full width on mobile */
        }
    }
    </style>
    ''' % (mime_type, bin_str)
    st.markdown(page_bg_img, unsafe_allow_html=True)

# Set background
bg_file = '466671893_2003287556785626_5199121047811111781_n.jpg'
if os.path.exists(bg_file):
    set_png_as_page_bg(bg_file)
else:
    set_png_as_page_bg('Screenshot_2.png')

st.title("🚀 Beyond Ads Scraper")

# -------------------------------------------------------
# Input Section
# -------------------------------------------------------
AREAS = [
    "Beirut", "Tripoli", "Sidon (Saida)", "Tyre (Sour)", "Jounieh", "Zahle", 
    "Nabatieh", "Baalbek", "Byblos (Jbeil)", "Batroun", "Aley", "Bhamdoun", "Broummana"
]
INDUSTRIES = [
    "Real Estate Companies", "Roofing Contractors", "Dentists", "Restaurants", 
    "Law Firms", "Hotels", "Hospitals", "Supermarkets", "Pharmacies", 
    "Schools", "Universities", "Gyms", "Car Rental", "Travel Agencies", "Banks", "Travel Agency"
]

# Only show inputs if not currently running to prevent changing params mid-scrape
main_placeholder = st.empty()

if not st.session_state.is_scraping and not st.session_state.results:
    with main_placeholder.form("scrape_form"):
        col1, col2 = st.columns(2)
        with col1:
            industry_input = st.selectbox("Select Industry", INDUSTRIES)
        with col2:
            area_input = st.selectbox("Select Area", AREAS)
        total_results = st.number_input("Max Results", min_value=1, max_value=500, value=5, step=1)
        
        submitted = st.form_submit_button("🚀 Start Scraping")

    if submitted:
        if industry_input and area_input:
            # Setup Scraper
            st.session_state.search_query = f"{industry_input} in {area_input}, Lebanon"
            st.session_state.results = []
            st.session_state.start_time = time.time()
            st.session_state.total_target = total_results
            
            # Filter logic
            strict_area_filter = None
            excluded_areas_list = []
            if "(" in area_input:
                strict_area_filter = area_input.split("(")[0].strip()
            else:
                strict_area_filter = area_input
            
            # Prepare Allowed Areas (Main Area + Sub Areas)
            allowed_areas_list = [strict_area_filter]
            if strict_area_filter in AREA_MAPPINGS:
                allowed_areas_list.extend(AREA_MAPPINGS[strict_area_filter])
            
            for area in AREAS:
                clean_area = area.split("(")[0].strip()
                if clean_area.lower() != strict_area_filter.lower():
                    excluded_areas_list.append(clean_area)

            # Start Scraper
            st.session_state.stop_event = threading.Event()
            st.session_state.results = []
            st.session_state.status_dict = {"text": "Starting...", "error": False}
            
            # Start background thread
            st.session_state.scraper_thread = threading.Thread(
                target=run_scraper_thread,
                args=(
                    st.session_state.search_query, 
                    total_results, 
                    strict_area_filter, 
                    excluded_areas_list,
                    allowed_areas_list,
                    st.session_state.results,
                    st.session_state.stop_event,
                    st.session_state.status_dict
                )
            )
            # Daemon thread ensures it dies if main process dies
            st.session_state.scraper_thread.daemon = True 
            st.session_state.scraper_thread.start()
            
            st.session_state.is_scraping = True
            main_placeholder.empty() # Clear the form
            
# -------------------------------------------------------
# Progress & Results Section
# -------------------------------------------------------
if st.session_state.is_scraping or st.session_state.results:
    st.info(f"Target: **{st.session_state.search_query}**")
    
    # Placeholders for dynamic content
    progress_bar = st.progress(0)
    metrics_placeholder = st.empty()
    status_placeholder = st.empty()
    dataframe_placeholder = st.empty()
    
    def render_metrics():
        current_count = len(st.session_state.results)
        target = st.session_state.total_target if 'total_target' in st.session_state else 1
        
        # Update Progress
        progress = min(int((current_count / target) * 100), 100)
        progress_bar.progress(progress)
        
        # Update Metrics
        m1, m2, m3 = metrics_placeholder.columns(3)
        m1.metric("Found", f"{current_count} / {target}")
        
        elapsed = time.time() - st.session_state.start_time
        m2.metric("Elapsed Time", f"{int(elapsed)}s")
        
        status_msg = "Running..." if st.session_state.is_scraping else "Completed"
        if st.session_state.status_dict.get("error"):
            status_msg = "Error"
        
        m3.metric("Status", status_msg)
        
        # Update DataFrame
        if st.session_state.results:
            df = pd.DataFrame(st.session_state.results)
            drop_cols = ["store_shipping", "in_store_pickup"]
            df = df.drop(columns=[c for c in drop_cols if c in df.columns], errors='ignore')
            
            # Sanitize DataFrame for Streamlit
            # We convert to string to ensure compatibility
            df = df.astype(str)
            
            # Use st.table (static HTML) for the last few rows to avoid Arrow/LargeUtf8 errors on frontend
            # This is a robust fallback since st.dataframe is crashing on Streamlit Cloud + Python 3.13
            try:
                # Show summary metric
                # Display only the last 10 results in a static HTML table to prevent crashes
                if len(df) > 0:
                    preview_cols = ["name", "address", "website", "phone_number", "instagram", "reviews_count"]
                    # Filter columns that exist in df
                    cols_to_show = [c for c in preview_cols if c in df.columns]
                    # If we have matches, use them, otherwise show all
                    if cols_to_show:
                        preview_df = df[cols_to_show].tail(10)
                    else:
                        preview_df = df.tail(10)

                    # Convert to HTML to avoid Arrow serialization entirely
                    html = preview_df.to_html(classes='dataframe', index=False)
                    # Add custom CSS to make it look decent
                    html = f"""
<style>
.dataframe {{
    font-family: sans-serif;
    border-collapse: collapse;
    width: 100%;
    background-color: #000000;
    color: #ffffff;
    font-size: 12px;
}}
.dataframe td, .dataframe th {{
    border: 1px solid #444;
    padding: 4px 6px;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    max-width: 200px;
}}
.dataframe tr:nth-child(even) {{background-color: #111;}}
.dataframe tr:hover {{background-color: #222;}}
.dataframe th {{
    padding-top: 8px;
    padding-bottom: 8px;
    text-align: left;
    background-color: #333;
    color: white;
}}
</style>
<h3>Live Preview (Last 10 results)</h3>
{html}
"""
                    dataframe_placeholder.markdown(html, unsafe_allow_html=True)
            except Exception as e:
                status_placeholder.warning(f"Could not render data table: {e}")
                
            return df
        return None

    # Initial Render
    df = render_metrics()
    
    # Scraping Loop (Blocking UI Update)
    if st.session_state.is_scraping:
        # Safety check: Ensure scraper_thread exists
        if "scraper_thread" not in st.session_state or st.session_state.scraper_thread is None:
            st.warning("Scraping state mismatch (Thread missing). Resetting.")
            st.session_state.is_scraping = False
            st.session_state.stop_event = None
            try:
                st.rerun()
            except AttributeError:
                st.experimental_rerun()
            except Exception:
                pass
            st.stop()

        stop_btn_placeholder = st.empty()
        if stop_btn_placeholder.button("⏹️ Stop Scraping (Hold to Stop)", type="primary"):
            st.session_state.stop_event.set()
            st.session_state.is_scraping = False
            try:
                st.rerun()
            except AttributeError:
                st.experimental_rerun()
            except Exception:
                pass

        # Robust loop with safety check
        while st.session_state.get("scraper_thread") and st.session_state.scraper_thread.is_alive():
            render_metrics()
            
            # Update detailed status
            if st.session_state.status_dict["text"]:
                status_placeholder.info(st.session_state.status_dict["text"])
            
            if st.session_state.status_dict["error"]:
                 st.error(st.session_state.status_dict["text"])
                 st.session_state.is_scraping = False
                 break
                 
            time.sleep(1)
            
        # Thread finished
        st.session_state.is_scraping = False
        render_metrics()
        
        if st.session_state.status_dict["error"]:
             st.error(f"Scraping failed: {st.session_state.status_dict['text']}")
        else:
             st.success("Scraping finished!")
             
        if 'stop_btn_placeholder' in locals():
            stop_btn_placeholder.empty()

    # Download & New Search (Only when finished)
    if not st.session_state.is_scraping and st.session_state.results:
        df = pd.DataFrame(st.session_state.results)
        # Re-clean for download
        drop_cols = ["store_shipping", "in_store_pickup"]
        df = df.drop(columns=[c for c in drop_cols if c in df.columns], errors='ignore')
        
        # Generate Excel
        safe_name = st.session_state.search_query.replace(" ", "_").replace(",", "").replace("/", "-")
        filename = f"Scrape_{safe_name}_{len(df)}.xlsx"
        
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # Make links clickable
            worksheet = writer.sheets['Sheet1']
            blue_font = Font(color="0563C1", underline="single")
            
            # Auto-adjust column widths
            for i, col in enumerate(df.columns):
                max_len = 0
                if col: max_len = len(str(col))
                for val in df[col]:
                    if val is not None: max_len = max(max_len, len(str(val)))
                adjusted_width = min(max_len + 2, 100) 
                col_letter = get_column_letter(i + 1)
                worksheet.column_dimensions[col_letter].width = adjusted_width

            cols_to_link = []
            if 'website' in df.columns: cols_to_link.append(df.columns.get_loc('website') + 1)
            if 'instagram' in df.columns: cols_to_link.append(df.columns.get_loc('instagram') + 1)
                
            for col_idx in cols_to_link:
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    if val and isinstance(val, str) and val.startswith("http"):
                        cell.hyperlink = val
                        cell.font = blue_font

        col_dl, col_new = st.columns([2, 1])
        with col_dl:
            st.download_button(
                label="📥 Download Excel Results",
                data=buffer.getvalue(),
                file_name=filename,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True
            )
        with col_new:
            if st.button("🔄 New Search", use_container_width=True):
                st.session_state.results = []
                st.session_state.search_query = ""
                st.session_state.is_scraping = False
                try: st.rerun()
                except: pass
